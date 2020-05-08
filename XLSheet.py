import openpyxl as xl
from openpyxl.chart import Reference, BarChart


def excelpro(filename):
    wb = xl.load_workbook(filename)
    sheetnum = input('Enter The Sheet Number to Work With ')
    print("Working on Sheet " + sheetnum)
    sheet = wb['Sheet' + sheetnum]
    for row in range(2, sheet.max_row + 1):
        cell=sheet.cell(row,3)
        correction = cell.value +10
        correct_cell = sheet.cell(row, 4)
        correct_cell.value = correction

    val = Reference(sheet, min_row=sheet.min_row+1, max_row=sheet.max_row, min_col=sheet.min_column+1, max_col=sheet.max_column)
    chart1 = BarChart()
    chart1.add_data(val)
    sheet.add_chart(chart1, 'a'+str(sheet.max_row+2))
    wb.save(filename)






